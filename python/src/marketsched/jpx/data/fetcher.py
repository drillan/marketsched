"""JPX data fetcher for SQ dates and holiday trading data.

This module provides functionality to fetch official data from JPX (Japan Exchange)
website and parse Excel files containing SQ dates and holiday trading information.

Example:
    >>> from marketsched.jpx.data.fetcher import JPXDataFetcher
    >>> fetcher = JPXDataFetcher()
    >>> sq_records = fetcher.fetch_sq_dates(years=[2026, 2027])
    >>> holiday_records = fetcher.fetch_holiday_trading()
"""

from datetime import date, datetime
from io import BytesIO

import httpx
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from marketsched.exceptions import DataFetchError, InvalidDataFormatError
from marketsched.jpx.data import HolidayTradingRecord, SQDateRecord

__all__ = ["JPXDataFetcher"]

# Expected column headers for validation
SQ_DATES_REQUIRED_COLUMNS = {"商品", "限月取引", "取引最終日", "権利行使日"}
HOLIDAY_TRADING_REQUIRED_COLUMNS = {"祝日取引の対象日", "名称", "実施有無"}

# Product name that has exercise date (SQ date)
SQ_DATE_PRODUCT = "日経225オプション"


class JPXDataFetcher:
    """Fetches market data from JPX official website.

    This class handles downloading Excel files from JPX and parsing them
    into structured data records.

    Attributes:
        timeout: HTTP request timeout in seconds.

    Example:
        >>> fetcher = JPXDataFetcher()
        >>> records = fetcher.fetch_sq_dates(years=[2026])
    """

    BASE_URL = "https://www.jpx.co.jp"
    SQ_DATES_PATH = "/derivatives/rules/last-trading-day/tvdivq0000004gz8-att/{year}_indexfutures_options_1_j.xlsx"
    HOLIDAY_TRADING_URL = "https://www.jpx.co.jp/derivatives/rules/holidaytrading/nlsgeu000006hweb-att/nlsgeu000006jgee.xlsx"

    def __init__(self, timeout: float = 30.0) -> None:
        """Initialize the fetcher.

        Args:
            timeout: HTTP request timeout in seconds.
        """
        self._timeout = timeout

    def fetch_sq_dates(self, years: list[int]) -> list[SQDateRecord]:
        """Fetch SQ date data for specified years.

        Downloads Excel files from JPX for each year and extracts SQ dates
        from 日経225オプション rows (which have exercise dates).

        Args:
            years: List of years to fetch data for.

        Returns:
            List of SQDateRecord objects sorted by contract month.

        Raises:
            DataFetchError: If network request fails.
            InvalidDataFormatError: If Excel format is unexpected.
        """
        all_records: list[SQDateRecord] = []

        for year in years:
            url = self.BASE_URL + self.SQ_DATES_PATH.format(year=year)
            excel_data = self._download_excel(url)
            records = self._parse_sq_dates_excel(excel_data)
            all_records.extend(records)

        if not all_records:
            raise InvalidDataFormatError("No SQ date records found in the data")

        return sorted(all_records, key=lambda r: r.contract_month)

    def fetch_holiday_trading(self) -> list[HolidayTradingRecord]:
        """Fetch holiday trading data.

        Downloads the holiday trading Excel file from JPX and extracts
        trading status for each holiday.

        Returns:
            List of HolidayTradingRecord objects sorted by date.

        Raises:
            DataFetchError: If network request fails.
            InvalidDataFormatError: If Excel format is unexpected.
        """
        excel_data = self._download_excel(self.HOLIDAY_TRADING_URL)
        records = self._parse_holiday_trading_excel(excel_data)

        return sorted(records, key=lambda r: r.date)

    def _download_excel(self, url: str) -> BytesIO:
        """Download an Excel file from URL.

        Args:
            url: URL to download from.

        Returns:
            BytesIO containing the Excel file content.

        Raises:
            DataFetchError: If download fails.
        """
        try:
            with httpx.Client(timeout=self._timeout) as client:
                response = client.get(url)
                response.raise_for_status()
                return BytesIO(response.content)
        except httpx.HTTPError as e:
            raise DataFetchError(url, str(e)) from e

    def _parse_sq_dates_excel(self, data: BytesIO) -> list[SQDateRecord]:
        """Parse SQ dates from Excel file.

        Args:
            data: BytesIO containing Excel file content.

        Returns:
            List of SQDateRecord objects.

        Raises:
            InvalidDataFormatError: If Excel format is unexpected.
        """
        wb = load_workbook(data, read_only=True)
        ws = wb.active
        if ws is None:
            raise InvalidDataFormatError("Excel file has no active worksheet")

        # Find header row and validate columns
        header_row_idx, column_map = self._find_sq_dates_header(ws)
        if header_row_idx is None:
            raise InvalidDataFormatError(
                f"Required columns not found: {SQ_DATES_REQUIRED_COLUMNS}"
            )

        records: list[SQDateRecord] = []

        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            product = row[column_map["商品"]]
            contract_month_raw = row[column_map["限月取引"]]
            last_trading_day_raw = row[column_map["取引最終日"]]
            exercise_date_raw = row[column_map["権利行使日"]]

            # Skip non-option products (they don't have exercise dates)
            if product != SQ_DATE_PRODUCT:
                continue

            # Skip rows without exercise date
            if exercise_date_raw is None or exercise_date_raw == "-":
                continue

            # Parse dates
            contract_month = self._parse_contract_month(contract_month_raw)
            if contract_month is None:
                continue

            last_trading_day = self._parse_date(last_trading_day_raw)
            sq_date = self._parse_date(exercise_date_raw)

            if last_trading_day is None or sq_date is None:
                continue

            records.append(
                SQDateRecord(
                    contract_month=contract_month,
                    last_trading_day=last_trading_day,
                    sq_date=sq_date,
                    product_category="index_futures_options",
                )
            )

        return records

    def _find_sq_dates_header(self, ws: Worksheet) -> tuple[int | None, dict[str, int]]:
        """Find the header row and map column indices.

        Args:
            ws: Excel worksheet.

        Returns:
            Tuple of (header_row_index, column_name_to_index_map).
            Returns (None, {}) if header not found.
        """
        for row_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True), 1):
            row_values = {str(v) if v else "" for v in row}

            if SQ_DATES_REQUIRED_COLUMNS.issubset(row_values):
                # Found header row, create column map
                column_map: dict[str, int] = {}
                for col_idx, value in enumerate(row):
                    if value and str(value) in SQ_DATES_REQUIRED_COLUMNS:
                        column_map[str(value)] = col_idx
                return row_idx, column_map

        return None, {}

    def _parse_holiday_trading_excel(self, data: BytesIO) -> list[HolidayTradingRecord]:
        """Parse holiday trading data from Excel file.

        Args:
            data: BytesIO containing Excel file content.

        Returns:
            List of HolidayTradingRecord objects.

        Raises:
            InvalidDataFormatError: If Excel format is unexpected.
        """
        wb = load_workbook(data, read_only=True)
        ws = wb.active
        if ws is None:
            raise InvalidDataFormatError("Excel file has no active worksheet")

        # Find header row and validate columns
        header_row_idx, column_map = self._find_holiday_header(ws)
        if header_row_idx is None:
            raise InvalidDataFormatError(
                f"Required columns not found: {HOLIDAY_TRADING_REQUIRED_COLUMNS}"
            )

        records: list[HolidayTradingRecord] = []

        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            holiday_date_raw = row[column_map["祝日取引の対象日"]]
            holiday_name = row[column_map["名称"]]
            status = row[column_map["実施有無"]]

            if holiday_date_raw is None or holiday_name is None:
                continue

            holiday_date = self._parse_date(holiday_date_raw)
            if holiday_date is None:
                continue

            is_trading = status == "実施する"

            records.append(
                HolidayTradingRecord(
                    date=holiday_date,
                    holiday_name=str(holiday_name),
                    is_trading=is_trading,
                    is_confirmed=True,  # Data from official source is confirmed
                )
            )

        return records

    def _find_holiday_header(self, ws: Worksheet) -> tuple[int | None, dict[str, int]]:
        """Find the holiday trading header row and map column indices.

        Args:
            ws: Excel worksheet.

        Returns:
            Tuple of (header_row_index, column_name_to_index_map).
            Returns (None, {}) if header not found.
        """
        for row_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True), 1):
            row_values = {str(v) if v else "" for v in row}

            if HOLIDAY_TRADING_REQUIRED_COLUMNS.issubset(row_values):
                # Found header row, create column map
                column_map: dict[str, int] = {}
                for col_idx, value in enumerate(row):
                    if value and str(value) in HOLIDAY_TRADING_REQUIRED_COLUMNS:
                        column_map[str(value)] = col_idx
                return row_idx, column_map

        return None, {}

    def _parse_contract_month(self, value: datetime | date | str | None) -> str | None:
        """Parse contract month to YYYYMM format.

        Args:
            value: Raw value from Excel (datetime, date, or string).

        Returns:
            Contract month in YYYYMM format, or None if parsing fails.
        """
        if value is None:
            return None

        if isinstance(value, datetime):
            return f"{value.year}{value.month:02d}"
        if isinstance(value, date):
            return f"{value.year}{value.month:02d}"

        return None

    def _parse_date(self, value: datetime | date | str | None) -> date | None:
        """Parse date from Excel cell value.

        Args:
            value: Raw value from Excel (datetime, date, or string).

        Returns:
            date object, or None if parsing fails.
        """
        if value is None:
            return None

        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str) and value == "-":
            return None

        return None
